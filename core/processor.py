import os
import sys
import logging
import pandas as pd
from datetime import datetime
import tempfile
from pathlib import Path
import json
import time
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import shutil
from PIL import Image as PILImage
import re
import io
from fpdf import FPDF

# Add parent directory to path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

# Import utils modules directly
from utils import config_manager
from utils import excel_utils
from utils import image_utils
import utils.image_utils as image_utils

# Import get_downloads_folder from config_manager
from utils.config_manager import get_downloads_folder

# Setup logging
logger = logging.getLogger(__name__)
# <<< Тест 1: Проверяем, работает ли логгер этого модуля >>>
logger.critical("--- Logger for core.processor initialized ---") 

# <<< Constants for image fitting >>>
DEFAULT_CELL_WIDTH_PX = 300  # Ширина ячейки по умолчанию в пикселях
DEFAULT_CELL_HEIGHT_PX = 120  # Высота ячейки по умолчанию в пикселях
DEFAULT_IMG_QUALITY = 90
MIN_IMG_QUALITY = 1  # Снижено с 5% до 1% для еще большего сжатия
MIN_KB_PER_IMAGE = 10
MAX_KB_PER_IMAGE = 2048 # 2MB max per image, prevents extreme cases
SIZE_BUDGET_FACTOR = 0.85 # Use 85% of total size budget for images
ROW_HEIGHT_PADDING = 1 # Минимальный отступ для высоты строки
MIN_ASPECT_RATIO = 0.5 # Минимальное соотношение сторон (высота/ширина)
MAX_ASPECT_RATIO = 2.0 # Максимальное соотношение сторон (высота/ширина)
EXCEL_WIDTH_TO_PIXEL_RATIO = 7.0  # Коэффициент преобразования единиц Excel в пиксели
EXCEL_PX_TO_PT_RATIO = 0.75  # Коэффициент преобразования пикселей в единицы Excel
DEFAULT_EXCEL_COLUMN_WIDTH = 40  # Ширина колонки в единицах Excel (примерно 300px)
MIN_COLUMN_WIDTH_PX = 100  # Минимальная допустимая ширина колонки в пикселях

# <<< Constants for progress formatting >>>
POWERSHELL_GREEN = '\033[92m'
POWERSHELL_YELLOW = '\033[93m'
POWERSHELL_CYAN = '\033[96m'
POWERSHELL_RESET = '\033[0m'
PROGRESS_SYMBOL = '■'

def print_progress(current, total, extra_info=""):
    """
    Выводит прогресс в PowerShell с цветным форматированием
    
    Args:
        current (int): Текущая позиция
        total (int): Общее количество элементов
        extra_info (str): Дополнительная информация для отображения
    """
    percent = round((current / total) * 100) if total > 0 else 0
    
    # Создаем цветную полосу прогресса
    bar_length = 30
    filled_length = int(bar_length * percent / 100)
    bar = (POWERSHELL_GREEN + PROGRESS_SYMBOL * filled_length + 
           POWERSHELL_YELLOW + '-' * (bar_length - filled_length) + 
           POWERSHELL_RESET)
    
    # Вычисляем длину текста прогресса без учета цветовых кодов
    progress_info = f"Прогресс: {bar} {percent}% ({current}/{total})"
    
    # Вычисляем точное количество пробелов для выравнивания
    box_width = 60
    # Вычитаем длину текста без учета цветовых ANSI кодов
    right_padding = box_width - len(f"Прогресс: ") - bar_length - len(f" {percent}% ({current}/{total})") - 1
    
    # Форматируем вывод с яркими цветами и точным выравниванием
    progress_text = f"\n{POWERSHELL_CYAN}╔{'═' * box_width}╗{POWERSHELL_RESET}"
    progress_text += f"\n{POWERSHELL_CYAN}║{POWERSHELL_RESET} {progress_info}{' ' * max(0, right_padding)}{POWERSHELL_CYAN}║{POWERSHELL_RESET}"
    
    if extra_info:
        # Правильное выравнивание для дополнительной информации
        info_padding = box_width - len(extra_info) - 1
        progress_text += f"\n{POWERSHELL_CYAN}║{POWERSHELL_RESET} {extra_info}{' ' * max(0, info_padding)}{POWERSHELL_CYAN}║{POWERSHELL_RESET}"
        
    progress_text += f"\n{POWERSHELL_CYAN}╚{'═' * box_width}╝{POWERSHELL_RESET}"
    
    print(progress_text, file=sys.stderr)
    sys.stderr.flush()

def ensure_temp_dir(prefix: str = "") -> str:
    """
    Создает и возвращает путь к временной директории.
    
    Args:
        prefix (str): Префикс для имени временной директории
    
    Returns:
        Путь к временной директории
    """
    temp_dir = os.path.join(tempfile.gettempdir(), f"{prefix}excelwithimages")
    os.makedirs(temp_dir, exist_ok=True)
    return temp_dir

def process_excel_file(
    file_path: str,
    article_col_name: str,
    image_folder: str,
    image_col_name: str = None,
    output_folder: str = None,
    max_total_file_size_mb: int = 100,
    progress_callback: callable = None,
    config: dict = None,
    header_row: int = 0,
    sheet_name: str = None,  # Добавляем параметр для имени листа
    secondary_image_folder: str = None,  # Папка с запасными изображениями (второй приоритет)
    tertiary_image_folder: str = None,   # Папка с дополнительными запасными изображениями (третий приоритет)
    output_filename: str = None,  # Имя выходного файла
    image_background_color: str = "000000"  # Цвет фона ячейки (по умолчанию черный)
) -> Tuple[str, Optional[pd.DataFrame], int, Dict[str, List[str]], List[str], List[Dict]]:
    """
    Обрабатывает Excel файл, вставляя изображения на основе номеров артикулов.
    
    Args:
        file_path (str): Путь к Excel файлу
        article_col_name (str): Имя столбца с артикулами (или буква столбца)
        image_folder (str): Путь к папке с изображениями
        image_col_name (str, optional): Имя столбца для вставки изображений (или буква). По умолчанию None
        output_folder (str, optional): Папка для сохранения результата. По умолчанию None
        max_total_file_size_mb (int, optional): Макс. размер файла в МБ. По умолчанию 100 МБ
        progress_callback (callable, optional): Функция для отображения прогресса. По умолчанию None
        config (dict, optional): Словарь с настройками. По умолчанию None
        header_row (int, optional): Номер строки заголовка (0-based). По умолчанию 0
        sheet_name (str, optional): Имя листа Excel для обработки. По умолчанию None (первый лист)
        secondary_image_folder (str, optional): Путь к папке с запасными изображениями. По умолчанию None
        tertiary_image_folder (str, optional): Путь к дополнительной папке с запасными изображениями. По умолчанию None
        output_filename (str, optional): Имя выходного файла. По умолчанию None
        image_background_color (str, optional): Цвет фона ячеек с изображениями в формате RRGGBB. По умолчанию "000000" (черный)
    
    Returns:
        Tuple[str, pd.DataFrame, int, Dict[str, List[str]], List[str], List[Dict]]: 
            - Путь к файлу результата
            - DataFrame с данными
            - Количество вставленных изображений
            - Словарь с артикулами, для которых найдено несколько изображений (ключ: артикул, значение: список путей)
            - Список артикулов, для которых не найдены изображения
            - Список результатов поиска изображений (словари с информацией о поиске)
    """
    # <<< Используем print в stderr вместо logger >>>
    print(">>> ENTERING process_excel_file <<<\n", file=sys.stderr)
    sys.stderr.flush()
    
    print(f"[PROCESSOR] Начало обработки: {file_path}", file=sys.stderr)
    print(f"[PROCESSOR] Параметры: article_col={article_col_name}, img_folder={image_folder}, img_col={image_col_name}, max_total_mb={max_total_file_size_mb}, sheet_name={sheet_name}", file=sys.stderr)

    # --- Валидация входных данных ---
    # Проверяем корректность обозначений колонок
    if not (article_col_name.isdigit() or article_col_name.isalpha()) or not (image_col_name.isdigit() or image_col_name.isalpha()):
        err_msg = f"Неверное обозначение колонки: '{article_col_name}' или '{image_col_name}'. Используйте буквенные (A, B, C...) или числовые (1, 2, 3...) обозначения"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)
        
    try:
        # Преобразуем числовые обозначения в буквенные
        if article_col_name.isdigit():
            article_col_idx = int(article_col_name)
            article_col_name = get_column_letter(article_col_idx)
            print(f"[PROCESSOR] Преобразовано числовое обозначение {article_col_idx} в букву {article_col_name}", file=sys.stderr)
            
        if image_col_name.isdigit():
            image_col_idx = int(image_col_name)
            image_col_name = get_column_letter(image_col_idx)
            print(f"[PROCESSOR] Преобразовано числовое обозначение {image_col_idx} в букву {image_col_name}", file=sys.stderr)
            
        article_col_idx = excel_utils.column_letter_to_index(article_col_name)
        image_col_idx = excel_utils.column_letter_to_index(image_col_name)
    except Exception as e:
        err_msg = f"Неверное обозначение колонки: '{article_col_name}' или '{image_col_name}'. Ошибка: {str(e)}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)
        
    if not os.path.exists(file_path):
        err_msg = f"Файл не найден: {file_path}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise FileNotFoundError(err_msg)
    
    if not os.path.exists(image_folder):
        err_msg = f"Папка с изображениями не найдена: {image_folder}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise FileNotFoundError(err_msg)

    # --- Чтение Excel ---
    try:
        # Если указан конкретный лист, читаем его
        if sheet_name:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                engine='openpyxl',
                skiprows=0,
                header=None
            )
            print(f"[PROCESSOR] Excel-файл прочитан в DataFrame (sheet={sheet_name}, header=None). Строк данных: {len(df)}", file=sys.stderr)
        else:
            df = pd.read_excel(file_path, header=0, engine='openpyxl') 
            print(f"[PROCESSOR] Excel-файл прочитан в DataFrame (header=0). Строк данных: {len(df)}", file=sys.stderr)
        
        # --- Загрузка книги openpyxl ---
        wb = openpyxl.load_workbook(file_path, read_only=False, keep_vba=False)
        try:
            # Проверяем наличие листов в книге
            if not wb.sheetnames:
                print("[PROCESSOR ERROR] В файле нет листов для обработки.", file=sys.stderr)
                raise ValueError("Excel-файл не содержит листов. Пожалуйста, выберите файл с данными.")
                
            # Фильтруем листы, исключая листы с макросами
            valid_sheets = [sheet_name for sheet_name in wb.sheetnames if not sheet_name.startswith('xl/macrosheets/')]
            if not valid_sheets:
                print("[PROCESSOR ERROR] В файле нет обычных листов, только макросы.", file=sys.stderr)
                raise ValueError("Внимание! Этот файл Excel содержит только макросы, а не обычные таблицы данных. Пожалуйста, выберите файл Excel с обычными листами, содержащими таблицы с артикулами и данными для обработки.")
            
            # Если указан лист, выбираем его, иначе используем активный
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"[PROCESSOR] Работаем с указанным листом: {sheet_name}", file=sys.stderr)
                else:
                    print(f"[PROCESSOR ERROR] Указанный лист {sheet_name} не найден в файле. Доступные листы: {wb.sheetnames}", file=sys.stderr)
                    raise ValueError(f"Лист '{sheet_name}' не найден в файле. Доступные листы: {wb.sheetnames}")
            else:
                # Используем первый лист
                ws = wb.active
                print(f"[PROCESSOR] Загружена рабочая книга, работаем с активным листом: {ws.title}", file=sys.stderr)
        except Exception as e:
            print(f"[PROCESSOR ERROR] Ошибка при выборе листа: {e}", file=sys.stderr)
            # Делаем сообщение об ошибке более понятным для пользователя
            if "'dict' object has no attribute 'shape'" in str(e):
                raise ValueError("Выбранный лист не содержит табличных данных. Пожалуйста, выберите лист с необходимыми данными.")
            else:
                raise ValueError(f"Ошибка при выборе листа: {e}")
        
    except Exception as e:
        err_msg = f"Ошибка при чтении Excel-файла: {e}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        # Выводим traceback в консоль
        import traceback
        traceback.print_exc(file=sys.stderr)
        
        # Делаем сообщение об ошибке более понятным для пользователя
        user_friendly_msg = err_msg
        if "'dict' object has no attribute 'shape'" in str(e):
            user_friendly_msg = "Выбранный лист не содержит табличных данных. Пожалуйста, выберите лист с необходимыми данными."
        elif "No sheet" in str(e) or "not found" in str(e):
            user_friendly_msg = "Указанный лист не найден в файле. Пожалуйста, выберите существующий лист."
        elif "Empty" in str(e) or "no data" in str(e):
            user_friendly_msg = "Выбранный лист не содержит данных. Пожалуйста, выберите лист с данными."
            
        raise RuntimeError(user_friendly_msg) from e

    if df.empty:
        err_msg = "Excel-файл не содержит данных"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)

    # --- Проверка существования колонки артикулов ---
    # Преобразуем букву колонки в индекс
    article_col_idx = excel_utils.column_letter_to_index(article_col_name)
    article_col_name = df.columns[article_col_idx]
    
    # Принудительно конвертируем значения артикулов в строковый тип
    df[article_col_name] = df[article_col_name].astype(str)
    
    articles = df[article_col_name].tolist()
    print(f"[PROCESSOR] Получено {len(articles)} артикулов из колонки {article_col_name}", file=sys.stderr)
    
    if article_col_name not in df.columns:
        err_msg = f"Колонка с артикулами '{article_col_name}' не найдена в файле. Доступные колонки: {list(df.columns)}"
        print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
        raise ValueError(err_msg)
    print(f"[PROCESSOR] Колонка с артикулами: '{article_col_name}'", file=sys.stderr)

    # --- Определение КОЛИЧЕСТВА строк с НЕНУЛЕВЫМИ артикулами для расчета лимита ---
    # Считаем строки, где артикул не пустой
    non_empty_article_rows = df[article_col_name].notna() & (df[article_col_name].astype(str).str.strip() != '')
    article_count = non_empty_article_rows.sum()
    
    if article_count == 0:
        article_count = 1 # Избегаем деления на ноль
        print("[PROCESSOR WARNING] Не найдено строк с непустыми артикулами для расчета лимита размера изображения. Используется значение по умолчанию.", file=sys.stderr)
    else:
        print(f"[PROCESSOR] Найдено {article_count} строк с непустыми артикулами.", file=sys.stderr)
        
    # --- Расчет лимита размера на одно изображение ---
    image_size_budget_mb = max_total_file_size_mb * SIZE_BUDGET_FACTOR
    target_kb_per_image = (image_size_budget_mb * 1024) / article_count if article_count > 0 else MAX_KB_PER_IMAGE
    target_kb_per_image = max(MIN_KB_PER_IMAGE, min(target_kb_per_image, MAX_KB_PER_IMAGE)) 
    print(f"[PROCESSOR] Расчетный лимит размера на изображение: {target_kb_per_image:.1f} КБ", file=sys.stderr)

    # --- Подготовка папки для обработанных изображений ---
    temp_image_dir_created = False
    if not image_folder:
        image_folder = ensure_temp_dir("processed_images_")
        temp_image_dir_created = True
        print(f"[PROCESSOR] Создана временная директория для обработанных изображений: {image_folder}", file=sys.stderr)
    elif not os.path.exists(image_folder):
         os.makedirs(image_folder)
         print(f"[PROCESSOR] Создана папка для обработанных изображений: {image_folder}", file=sys.stderr)


    # --- Подготовка к вставке изображений ---
    try:
        # НАПРЯМУЮ ИСПОЛЬЗУЕМ УКАЗАННУЮ БУКВУ КОЛОНКИ
        image_col_letter_excel = image_col_name
        print(f"[PROCESSOR] Изображения будут вставляться в колонку: '{image_col_letter_excel}'", file=sys.stderr)
    except Exception as e:
         err_msg = f"Ошибка при подготовке колонки для изображений ('{article_col_name}'): {e}"
         print(f"[PROCESSOR ERROR] {err_msg}", file=sys.stderr)
         import traceback
         traceback.print_exc(file=sys.stderr)
         raise RuntimeError(err_msg) from e

    # --- Настройка ШИРИНЫ КОЛОНКИ ---
    try:
        # ВАЖНО: НЕ устанавливаем ширину колонки вручную, используем текущую ширину Excel
        # Проверяем фактическую ширину колонки
        column_width_excel = ws.column_dimensions[image_col_letter_excel].width
        
        # Если ширина колонки не определена, используем стандартную ширину Excel (8.43)
        if not column_width_excel:
            column_width_excel = ws.sheet_format.defaultColWidth or 8.43
        
        # Переводим в пиксели для информации
        actual_width_px = int(column_width_excel * EXCEL_WIDTH_TO_PIXEL_RATIO)
        print(f"[PROCESSOR] Фактическая ширина столбца {image_col_letter_excel}: {column_width_excel:.2f} ед. Excel (≈ {actual_width_px} пикс.)", file=sys.stderr)
    except Exception as e:
        print(f"[PROCESSOR WARNING] Не удалось определить ширину столбца {image_col_letter_excel}: {e}", file=sys.stderr)

    # --- Обработка строк и вставка изображений ---
    images_inserted = 0
    rows_processed = 0
    total_processed_image_size_kb = 0
    
    # Создаем список для хранения результатов поиска изображений
    image_search_results = []
    
    # Инициализируем списки для хранения результатов
    not_found_articles = []
    multiple_images_found = {}
    
    print("[PROCESSOR] --- Начало итерации по строкам DataFrame ---", file=sys.stderr)
    
    # Сбрасываем кеш качества перед обработкой нового файла
    from utils.image_utils import cached_quality
    image_utils.cached_quality = None
    print("[PROCESSOR] Кеш качества изображений сброшен", file=sys.stderr)
    
    # Переменные для определения оптимального качества сжатия
    successful_quality = DEFAULT_IMG_QUALITY  # Если не найдено, используем значение по умолчанию
    quality_determined = False  # Флаг, указывающий, был ли определен уровень качества
    
    # Общее количество строк для расчета прогресса
    total_rows = len(df)
    
    # Итерация по строкам таблицы
    for excel_row_index, row in df.iterrows():
        # Проверяем, нужно ли обновить прогресс
        if progress_callback and excel_row_index % 5 == 0:  # Обновление каждые 5 строк
            progress_value = min(0.9, (excel_row_index / len(df)) * 0.9)  # 90% прогресса на обработку строк
            progress_callback(progress_value, f"Обработка строки {excel_row_index + 1} из {len(df)}")
        
        rows_processed += 1
        
        # Сначала получаем артикул
        article_str = str(row[article_col_name]).strip()
        
        print(f"[PROCESSOR] Обработка строки {excel_row_index}, артикул: '{article_str}'", file=sys.stderr)
        
        if pd.isna(row[article_col_name]) or article_str.strip() == "":
            print(f"[PROCESSOR]   Пустой артикул в строке {excel_row_index}, пропускаем", file=sys.stderr)
            continue
        
        # Find images for this article in multiple folders
        supported_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')
        
        # Получаем пути к резервным папкам из параметров или конфигурации
        secondary_folder_path = secondary_image_folder or config_manager.get_setting("paths.secondary_images_folder_path", "")
        tertiary_folder_path = tertiary_image_folder or config_manager.get_setting("paths.tertiary_images_folder_path", "")
        
        # Логируем папки для диагностики
        print(f"[PROCESSOR DEBUG] Поиск изображений для артикула '{article_str}' в папках:", file=sys.stderr)
        print(f"[PROCESSOR DEBUG]   Основная: {image_folder}", file=sys.stderr)
        print(f"[PROCESSOR DEBUG]   Вторичная: {secondary_folder_path}", file=sys.stderr)
        print(f"[PROCESSOR DEBUG]   Третичная: {tertiary_folder_path}", file=sys.stderr)
        
        search_result = image_utils.find_images_in_multiple_folders(
            article_str, 
            image_folder,
            secondary_folder_path,
            tertiary_folder_path,
            supported_extensions,
            search_recursively=True
        )
        
        # Добавляем результат поиска в список
        search_result['row_index'] = excel_row_index
        search_result['article'] = article_str
        search_result['image_folders'] = {
            'primary': image_folder,
            'secondary': secondary_folder_path,
            'tertiary': tertiary_folder_path
        }
        image_search_results.append(search_result)
        
        # If no images found, record and continue
        if not search_result["found"]:
            print(f"[PROCESSOR WARNING]   Для артикула '{article_str}' (строка {excel_row_index}) не найдено изображений. Пропускаем.", file=sys.stderr)
            # Добавляем артикул в список не найденных
            not_found_articles.append(article_str)
            continue
        
        # If multiple images found, record for report
        all_image_paths = search_result["images"]
        source_folder_priority = search_result["source_folder"]
        
        if len(all_image_paths) > 1:
            print(f"[PROCESSOR INFO]   Найдено несколько изображений для артикула '{article_str}': {len(all_image_paths)}", file=sys.stderr)
            multiple_images_found[article_str] = all_image_paths
            # Still proceed with the first image
        
        image_path = all_image_paths[0]
        print(f"[PROCESSOR]   Выбрано первое найденное изображение: {image_path} (папка приоритета {source_folder_priority})", file=sys.stderr)

        # Проверяем, удовлетворяет ли изображение требованиям по размеру
        original_size_kb = os.path.getsize(image_path) / 1024
        print(f"[PROCESSOR]   Размер исходного изображения: {original_size_kb:.1f} КБ, лимит: {target_kb_per_image:.1f} КБ", file=sys.stderr)
        
        # 1. ОПТИМИЗАЦИЯ ИЗОБРАЖЕНИЯ (если требуется)
        optimized_buffer = None
        
        if original_size_kb <= target_kb_per_image:
            # Если размер уже подходит, просто загружаем изображение без оптимизации
            print(f"[PROCESSOR]   Изображение уже удовлетворяет требованиям по размеру, загружаем без оптимизации", file=sys.stderr)
            try:
                with open(image_path, 'rb') as f_orig:
                    optimized_buffer = io.BytesIO(f_orig.read())
                print(f"[PROCESSOR]   Загружено без оптимизации, размер: {optimized_buffer.tell()/1024:.1f} КБ", file=sys.stderr)
                optimized_buffer.seek(0)
            except Exception as e:
                print(f"[PROCESSOR ERROR]   Ошибка при загрузке изображения без оптимизации: {e}", file=sys.stderr)
                # Если не удалось загрузить, попробуем оптимизировать
        else:
            # Требуется оптимизация
            print(f"[PROCESSOR]   Вызов optimize_image_for_excel для {image_path} с лимитом {target_kb_per_image:.1f} КБ", file=sys.stderr)
            
            try:
                # Если уровень качества еще не определен, определяем его на первом изображении,
                # которое требует оптимизации
                if not quality_determined:
                    # Ищем оптимальное качество для сжатия
                    print(f"[PROCESSOR]   ОПРЕДЕЛЕНИЕ ОПТИМАЛЬНОГО КАЧЕСТВА: поиск качества от {DEFAULT_IMG_QUALITY}% до {MIN_IMG_QUALITY}%", file=sys.stderr)
                    optimized_buffer = image_utils.optimize_image_for_excel(
                        image_path, 
                        target_size_kb=target_kb_per_image,
                        quality=DEFAULT_IMG_QUALITY,
                        min_quality=MIN_IMG_QUALITY    
                    )
                    
                    # Помечаем что определили качество
                    quality_determined = True
                else:
                    # Для всех последующих изображений используем кешированное качество
                    print(f"[PROCESSOR]   Используем кешированное качество для изображения {image_path}", file=sys.stderr)
                    optimized_buffer = image_utils.optimize_image_for_excel(
                        image_path, 
                        target_size_kb=target_kb_per_image,
                        quality=DEFAULT_IMG_QUALITY,
                        min_quality=MIN_IMG_QUALITY    
                    )
            except Exception as e:
                print(f"[PROCESSOR ERROR]   Ошибка при оптимизации изображения: {e}", file=sys.stderr)
                # Если не удалось оптимизировать, попробуем загрузить оригинальное изображение
                try:
                    with open(image_path, 'rb') as f_orig:
                        optimized_buffer = io.BytesIO(f_orig.read())
                    print(f"[PROCESSOR]   Загружен оригинал из-за ошибки оптимизации, размер: {optimized_buffer.tell()/1024:.1f} КБ", file=sys.stderr)
                    optimized_buffer.seek(0)
                except Exception as load_e:
                    print(f"[PROCESSOR ERROR]   Не удалось загрузить оригинальное изображение: {load_e}", file=sys.stderr)
                    continue
        
        if optimized_buffer and optimized_buffer.getbuffer().nbytes > 0:
            buffer_size_kb = optimized_buffer.tell() / 1024
            print(f"[PROCESSOR]   Размер буфера для вставки: {buffer_size_kb:.1f} КБ", file=sys.stderr)
            current_image_size_kb = buffer_size_kb
            total_processed_image_size_kb += current_image_size_kb
            
            # Дополнительная проверка буфера - убеждаемся, что это действительно изображение
            optimized_buffer.seek(0)
            try:
                verification_img = PILImage.open(optimized_buffer)
                img_format = verification_img.format
                img_width_px, img_height_px = verification_img.size
                print(f"[PROCESSOR]   ВЕРИФИКАЦИЯ: буфер содержит изображение формата {img_format}, {img_width_px}x{img_height_px}", file=sys.stderr)
                
                # Создаем временную копию буфера для сохранения в файл (для отладки)
                try:
                    debug_copy = io.BytesIO(optimized_buffer.getvalue())
                    temp_debug_path = os.path.join(tempfile.gettempdir(), f"debug_image_{time.time()}.jpg")
                    with open(temp_debug_path, "wb") as debug_file:
                        debug_file.write(debug_copy.getvalue())
                    print(f"[PROCESSOR]   Создана отладочная копия изображения: {temp_debug_path}", file=sys.stderr)
                except Exception as debug_e:
                    print(f"[PROCESSOR]   Примечание: не удалось создать отладочную копию: {debug_e}", file=sys.stderr)
                
                # Сбрасываем указатель в начало буфера после верификации
                optimized_buffer.seek(0)
            except Exception as verify_e:
                print(f"[PROCESSOR] ОШИБКА ВЕРИФИКАЦИИ: Буфер не содержит корректного изображения: {verify_e}", file=sys.stderr)
                # Пробуем сохранить проблемный буфер для анализа
                try:
                    error_path = os.path.join(tempfile.gettempdir(), f"error_buffer_{time.time()}.bin")
                    with open(error_path, "wb") as error_file:
                        error_file.write(optimized_buffer.getvalue())
                    print(f"[PROCESSOR]   Сохранён проблемный буфер для анализа: {error_path}", file=sys.stderr)
                except Exception as err_save_e:
                    print(f"[PROCESSOR]   Не удалось сохранить проблемный буфер: {err_save_e}", file=sys.stderr)
                    
                # Если буфер некорректен, пробуем загрузить оригинальное изображение
                try:
                    print(f"[PROCESSOR]   Пробуем загрузить оригинальное изображение как резервный вариант", file=sys.stderr)
                    with open(image_path, "rb") as original_file:
                        optimized_buffer = io.BytesIO(original_file.read())
                    print(f"[PROCESSOR]   Загружено оригинальное изображение размером {optimized_buffer.getbuffer().nbytes / 1024:.1f} КБ", file=sys.stderr)
                    optimized_buffer.seek(0)
                    verification_img = PILImage.open(optimized_buffer)
                    img_width_px, img_height_px = verification_img.size
                except Exception as orig_load_e:
                    print(f"[PROCESSOR] КРИТИЧЕСКАЯ ОШИБКА: Не удалось загрузить даже оригинальное изображение: {orig_load_e}", file=sys.stderr)
                    continue  # Пропускаем эту итерацию
            
            # Получаем размеры изображения напрямую из буфера
            try:
                optimized_buffer.seek(0)
                img = PILImage.open(optimized_buffer)
                img_width_px, img_height_px = img.size
                print(f"[PROCESSOR]     Получены размеры из буфера: {img_width_px}x{img_height_px}", file=sys.stderr)
            except Exception as dim_e:
                print(f"[PROCESSOR] WARNING: Не удалось получить размеры изображения из буфера: {dim_e}", file=sys.stderr)
            
            # Вставляем изображение в Excel
            try:
                # Проверяем, что буфер изображения не пустой
                if not optimized_buffer or optimized_buffer.getbuffer().nbytes == 0:
                    print(f"[PROCESSOR WARNING] Пустой буфер изображения для артикула '{article_str}' (строка {excel_row_index})", file=sys.stderr)
                    continue
                
                # 1. Определяем фактическую ширину колонки Excel
                column_width_excel = None
                try:
                    # Получаем прямой доступ к размеру колонки
                    column_width_excel = ws.column_dimensions[image_col_letter_excel].width
                except Exception:
                    pass
                
                # Если ширина не определена, используем стандартную ширину листа
                if not column_width_excel:
                    column_width_excel = ws.sheet_format.defaultColWidth or 8.43  # Стандартный размер колонки Excel
                
                # Переводим в пиксели для расчета размеров изображения
                target_width_px = int(column_width_excel * EXCEL_WIDTH_TO_PIXEL_RATIO)
                print(f"[PROCESSOR] Используем фактическую ширину столбца {image_col_letter_excel}: {column_width_excel:.2f} ед. Excel ({target_width_px} пикс.)", file=sys.stderr)
                
                # Убираем корректировку - используем точную ширину столбца
                # 2. Получаем размеры исходного изображения для сохранения пропорций
                optimized_buffer.seek(0)
                pil_image = PILImage.open(optimized_buffer)
                img_width, img_height = pil_image.size
                aspect_ratio = img_height / img_width if img_width > 0 else 1.0
                optimized_buffer.seek(0)
                print(f"[PROCESSOR] Размеры оригинального изображения: {img_width}x{img_height}, соотношение сторон: {aspect_ratio:.2f}", file=sys.stderr)
                
                # Рассчитываем высоту изображения с сохранением пропорций
                target_height_px = int(target_width_px * aspect_ratio)
                
                # Формируем адрес ячейки для вставки
                anchor_cell = f"{image_col_letter_excel}{excel_row_index + 1 + header_row}"
                
                # Вставляем изображение с рассчитанными размерами и черным фоном
                print(f"[PROCESSOR] Вставляем изображение с размерами: {target_width_px}x{target_height_px} пикс. и черным фоном", file=sys.stderr)
                excel_utils.insert_image_from_buffer(
                    ws, 
                    optimized_buffer,
                    anchor_cell,
                    width=target_width_px,
                    height=target_height_px,
                    preserve_aspect_ratio=True,
                    background_color=image_background_color  # Черный фон
                )
                
                # 3. Устанавливаем высоту строки, чтобы изображение точно вписалось
                row_num = excel_row_index + 1 + header_row
                # Преобразуем пиксели в единицы Excel и добавляем 1 пиксель к высоте
                row_height_excel = (target_height_px + 1) * EXCEL_PX_TO_PT_RATIO
                excel_utils.set_row_height(ws, row_num, row_height_excel)
                print(f"[PROCESSOR] Установлена высота строки {row_num}: {row_height_excel:.2f} ед. Excel для вмещения изображения (с запасом +1px)", file=sys.stderr)
                
                # Увеличиваем счетчик успешно вставленных изображений
                images_inserted += 1
                print(f"[PROCESSOR] Изображение успешно вставлено в ячейку {anchor_cell}", file=sys.stderr)
                
            except Exception as e:
                print(f"[PROCESSOR ERROR] Ошибка при вставке изображения: {e}", file=sys.stderr)
                traceback.print_exc(file=sys.stderr)
                # Если количество вставленных изображений > 0, продолжаем
                if images_inserted > 0:
                    print(f"[PROCESSOR WARNING] Вставка изображения не удалась, но продолжаем обработку других строк", file=sys.stderr)
                    continue
                else:
                    # Это первое изображение и мы получили ошибку
                    print(f"[PROCESSOR ERROR] Критическая ошибка при вставке первого изображения: {e}", file=sys.stderr)
                    raise
        else:
            print(f"[PROCESSOR WARNING] Пустой буфер изображения для артикула '{article_str}' (строка {excel_row_index})", file=sys.stderr)
        
        # Вывод прогресса обработки строк в процентах
        extra_info = f"Строка: {excel_row_index + 1}, артикул: {article_str}"
        print_progress(rows_processed, total_rows, extra_info)
    
    # --- Сохранение результата ---
    print("\n[PROCESSOR] --- Сохранение результата ---", file=sys.stderr)
    
    try:
        # Создаем папку для результатов, если не существует
        if not output_folder:
            output_folder = os.path.join(os.path.dirname(file_path), "processed")
        
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
            print(f"[PROCESSOR] Создана папка для результатов: {output_folder}", file=sys.stderr)
        
        # Генерируем уникальное имя файла с датой и временем
        if output_filename:
            result_file_path = os.path.join(output_folder, output_filename)
        else:
            output_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_with Images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            result_file_path = os.path.join(output_folder, output_filename)
        
        # Сохраняем Excel-файл
        try:
            wb.save(result_file_path)
            print(f"[PROCESSOR] Результат сохранен в файл: {result_file_path}", file=sys.stderr)
            
            # Получаем фактический размер файла
            file_size_mb = os.path.getsize(result_file_path) / (1024 * 1024)
            print(f"[PROCESSOR] Фактический размер файла: {file_size_mb:.2f} МБ", file=sys.stderr)
            
            if progress_callback:
                progress_callback(1.0, f"Готово. Размер файла: {file_size_mb:.2f} MB")
        except Exception as save_e:
            print(f"[PROCESSOR] ОШИБКА ПРИ СОХРАНЕНИИ EXCEL: {save_e}", file=sys.stderr)
            # Вывод подробной ошибки в лог
            import traceback
            traceback.print_exc(file=sys.stderr)
            raise RuntimeError(f"Ошибка при сохранении файла: {save_e}")
    except Exception as out_e:
        print(f"[PROCESSOR] ОШИБКА ПРИ ПОДГОТОВКЕ ВЫВОДА: {out_e}", file=sys.stderr)
        raise RuntimeError(f"Ошибка при подготовке вывода: {out_e}")
    
    print(f"[PROCESSOR] СТАТИСТИКА: Обработано строк: {rows_processed}, вставлено изображений: {images_inserted}", file=sys.stderr)
    print(f"[PROCESSOR] Общий размер вставленных изображений: {total_processed_image_size_kb:.2f} КБ", file=sys.stderr)
    
    # Финальный вывод прогресса обработки
    print_progress(total_rows, total_rows, f"Завершено! Вставлено изображений: {images_inserted}")
    
    # Добавляем результаты поиска изображений к возвращаемым данным
    return result_file_path, df, images_inserted, multiple_images_found, not_found_articles, image_search_results

def get_column_width_pixels(ws, column_letter):
    """
    Получает фактическую ширину колонки в пикселях на основе настроек Excel.
    Не использует значений по умолчанию, берет только реальные данные из Excel.
    
    Args:
        ws: Рабочий лист Excel
        column_letter: Буква колонки (например, 'A', 'B', etc.)
        
    Returns:
        int: Ширина колонки в пикселях
    """
    try:
        # Получаем размер колонки из объекта column_dimensions
        column_dimensions = ws.column_dimensions.get(column_letter)
        
        # Проверяем, существует ли размер для данной колонки
        if column_dimensions and hasattr(column_dimensions, 'width') and column_dimensions.width is not None:
            width_in_excel_units = column_dimensions.width
            print(f"[PROCESSOR DEBUG] Получена ширина колонки {column_letter}: {width_in_excel_units} ед. Excel", file=sys.stderr)
        else:
            # Используем стандартную ширину из настроек листа
            width_in_excel_units = ws.sheet_format.defaultColWidth or 8.43  # Стандартный размер колонки Excel
            print(f"[PROCESSOR DEBUG] Используется стандартная ширина листа для колонки {column_letter}: {width_in_excel_units} ед. Excel", file=sys.stderr)
        
        # Преобразуем единицы Excel в пиксели
        pixels = int(width_in_excel_units * EXCEL_WIDTH_TO_PIXEL_RATIO)
        print(f"[PROCESSOR DEBUG] Ширина колонки {column_letter} в пикселях: {pixels} px", file=sys.stderr)
        return pixels
    except Exception as e:
        print(f"[PROCESSOR WARNING] Ошибка при получении ширины колонки {column_letter}: {e}", file=sys.stderr)
        # Используем стандартную ширину Excel в крайнем случае
        standard_width = 8.43  # Стандартная ширина колонки Excel
        return int(standard_width * EXCEL_WIDTH_TO_PIXEL_RATIO)

def _get_col_index(col_identifier: str, df_columns: pd.Index) -> int:
    """
    Преобразует идентификатор столбца (букву или номер) в 0-индексный номер.
    """
    if isinstance(col_identifier, int):
        return col_identifier

    if col_identifier.isdigit():
        # 1-based index from user
        col_idx = int(col_identifier) - 1
        if 0 <= col_idx < len(df_columns):
            return col_idx
        else:
            raise ValueError(f"Номер столбца '{col_identifier}' выходит за пределы таблицы (доступно столбцов: {len(df_columns)}).")
    elif col_identifier.isalpha():
        # Letter from user
        try:
            # openpyxl is 1-based
            col_idx = column_index_from_string(col_identifier.upper()) - 1
            if 0 <= col_idx < len(df_columns):
                return col_idx
            else:
                raise ValueError(f"Столбец '{col_identifier}' выходит за пределы таблицы (доступно столбцов: {len(df_columns)}).")
        except ValueError:
             raise ValueError(f"Некорректная буква столбца: '{col_identifier}'.")
    else:
        # Try to find by name - this won't work with header=None but it's a good fallback
        if col_identifier in df_columns:
            return df_columns.get_loc(col_identifier)
        else:
            raise ValueError(f"Столбец '{col_identifier}' не найден.")

def find_image_path(article: str, folders: List[str]) -> Optional[str]:
    """
    Ищет изображение по артикулу в списке папок, включая подпапки (рекурсивно).
    Логика поиска соответствует оригинальной программе.
    """
    logger.debug(f"Поиск для артикула '{article}' в папках: {folders}")
    supported_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')

    for folder in folders:
        if not folder or not os.path.exists(folder):
            if folder:
                logger.warning(f"Папка с изображениями не существует или недоступна: {folder}")
            continue
        
        for root, _, files in os.walk(folder):
            for file in files:
                file_name_without_ext, extension = os.path.splitext(file)
                if extension.lower() in supported_extensions:
                    if file_name_without_ext == article:
                        img_path = os.path.join(root, file)
                        logger.info(f"Найдено изображение для артикула '{article}': {img_path}")
                        return img_path

    logger.warning(f"Изображение для артикула '{article}' не найдено ни в одной из папок.")
    return None

def _force_wrap_text(pdf: FPDF, text: str, max_width: float) -> str:
    """
    A robust text processing function that prepares text for FPDF's multi_cell.
    It ensures text fits within the cell width without breaking words across lines.
    1. Replaces any single character that is too wide with '?'.
    2. Keeps words intact for dynamic font sizing in create_pdf_cards.
    """
    safe_words = []
    # Replace non-breaking spaces and strip text as a precaution
    text = text.replace('\u00A0', ' ').strip()

    for word in text.split(' '):
        if not word:
            continue
        
        # 1. Sanitize the word of any single character that is too wide
        sanitized_word = ""
        for char in word:
            if pdf.get_string_width(char) > max_width:
                logger.warning(f"A single character ('{char}') was wider than the cell and has been replaced by '?'.") 
                sanitized_word += "?"
            else:
                sanitized_word += char
        
        # 2. Keep words intact for dynamic font sizing
        safe_words.append(sanitized_word)
            
    return " ".join(safe_words)

def create_pdf_cards(
    df: pd.DataFrame,
    article_col_name: str,
    product_image_folders: List[str],
    package_image_folders: List[str],
    output_folder: str,
    progress_callback: callable = None,
    max_total_file_size_mb: int = 100,
) -> Tuple[str, int, List[str]]:
    """
    Создает PDF-файл с карточками товаров.
    """
    image_utils.cached_quality = None # Reset cached quality for each new processing session

    pdf = FPDF(orientation='P', unit='mm', format=(90, 160))
    
    # Используем шрифт Arial, который стандартно установлен в Windows
    font_family = 'Arial'
    try:
        # Добавляем обычный шрифт
        pdf.add_font(font_family, '', 'C:/Windows/Fonts/arial.ttf', uni=True)
        # Добавляем жирный шрифт
        pdf.add_font(font_family, 'B', 'C:/Windows/Fonts/arialbd.ttf', uni=True)
        pdf.set_font(font_family, '', 14)
    except RuntimeError as e:
        logger.warning(f"Не удалось загрузить шрифт Arial: {e}. Используется стандартный шрифт, кириллица может не отображаться.")
        # Fallback to a core font, which may not render Cyrillic correctly
        font_family = 'Helvetica'
        pdf.set_font(font_family, '', 14)

    inserted_cards = 0
    not_found_articles = []
    
    try:
        article_col_idx = _get_col_index(article_col_name, df.columns)
    except ValueError as e:
        # Re-raise with a more user-friendly message
        raise ValueError(f"Ошибка в указании столбца с артикулами: {e}")

    # Получаем заголовки из первой строки DataFrame
    headers = df.iloc[0].tolist() if len(df) > 0 else []
    
    # Пропускаем первую строку (заголовки) и обрабатываем только данные
    data_df = df.iloc[1:] if len(df) > 1 else pd.DataFrame()
    
    if data_df.empty:
        logger.warning("После пропуска строки с заголовками не осталось данных для обработки")
        return "", 0, not_found_articles

    total_rows = len(data_df)
    for index, row in data_df.iterrows():
        if progress_callback:
            progress_callback(index - 1 + 1, total_rows)  # Корректируем индекс, так как пропустили первую строку

        try:
            article = str(row.iloc[article_col_idx]).strip()
        except IndexError:
            # This case should be caught by _get_col_index, but as a safeguard:
            raise IndexError(f"Столбец с артикулами ({article_col_name}) не существует в файле.")

        product_img_path = find_image_path(article, product_image_folders)
        package_img_path = find_image_path(article, package_image_folders)

        # Рассчитываем лимит размера на изображение
        article_count = len(data_df)
        if article_count == 0:
            article_count = 1  # Избегаем деления на ноль
            
        image_size_budget_mb = max_total_file_size_mb * SIZE_BUDGET_FACTOR
        target_kb_per_image = (image_size_budget_mb * 1024) / article_count if article_count > 0 else MAX_KB_PER_IMAGE
        target_kb_per_image = max(MIN_KB_PER_IMAGE, min(target_kb_per_image, MAX_KB_PER_IMAGE))
        
        logger.debug(f"Лимит размера на изображение: {target_kb_per_image:.1f} КБ")

        # Создаем страницу для каждого артикула, даже если изображения отсутствуют
        pdf.add_page()
        
        # Если хотя бы одно изображение отсутствует, добавляем артикул в список "ненайденных"
        if not product_img_path or not package_img_path:
            not_found_articles.append(article)

        # Добавляем изображения, только если они были найдены
        if product_img_path:
            try:
                # Оптимизируем изображение перед вставкой
                optimized_buffer = image_utils.optimize_image_for_excel(
                    product_img_path,
                    target_size_kb=target_kb_per_image,
                    quality=DEFAULT_IMG_QUALITY,
                    min_quality=MIN_IMG_QUALITY
                )
                temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_product_{article}.jpg")
                with open(temp_img_path, "wb") as f:
                    f.write(optimized_buffer.getvalue())
                pdf.image(temp_img_path, x=5, y=5, w=40)
                os.remove(temp_img_path)
            except Exception as e:
                logger.error(f"Ошибка при вставке изображения товара '{product_img_path}' для артикула '{article}': {e}")
        if package_img_path:
            try:
                # Оптимизируем изображение перед вставкой
                optimized_buffer = image_utils.optimize_image_for_excel(
                    package_img_path,
                    target_size_kb=target_kb_per_image,
                    quality=DEFAULT_IMG_QUALITY,
                    min_quality=MIN_IMG_QUALITY
                )
                temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_package_{article}.jpg")
                with open(temp_img_path, "wb") as f:
                    f.write(optimized_buffer.getvalue())
                pdf.image(temp_img_path, x=45, y=5, w=40)
                os.remove(temp_img_path)
            except Exception as e:
                logger.error(f"Ошибка при вставке изображения упаковки '{package_img_path}' для артикула '{article}': {e}")


        # Add text
        pdf.set_y(50)
        
        # Собираем текст из всех ячеек строки, включая артикул в его исходном порядке
        text_lines = []
        # Используем заголовки из первой строки и значения из текущей строки
        for i in range(len(row)):
            cell_value = str(row.iloc[i]).strip()
            if cell_value and cell_value.lower() != 'nan':
                # Получаем заголовок для текущей колонки
                header = headers[i] if i < len(headers) else f"Столбец {i+1}"
                header = str(header).strip()
                if header and header.lower() != 'nan':
                    # Добавляем заголовок и значение как отдельные элементы для таблицы
                    text_lines.append({"header": header, "value": cell_value})

        # --- Dynamically adjust font size and column width ---
        available_width = pdf.w - pdf.l_margin - pdf.r_margin
        available_height = 155 - pdf.y 

        best_font_size = 0
        final_lines_to_render = []

        # Функция для проверки, помещается ли текст в одну строку
        def text_fits_in_one_line(font_size, text, max_width):
            pdf.set_font_size(font_size)
            return pdf.get_string_width(text) <= max_width

        # Определяем максимальную ширину заголовка для динамического расчета ширины колонок
        max_header_width = 0
        for item in text_lines:
            pdf.set_font(font_family, 'B', 14)  # Используем максимальный размер шрифта для оценки
            header_width = pdf.get_string_width(item['header'])
            max_header_width = max(max_header_width, header_width)

        # Ограничиваем максимальную ширину заголовка до 40% от доступной ширины
        max_header_width = min(max_header_width + 10, available_width * 0.4)  # +10 для отступа
        value_width = available_width - max_header_width - 10  # -10 для отступа между колонками

        # Iterate from a reasonable max down to a min font size to find the best fit
        for test_font_size in range(14, 5, -1):
            pdf.set_font_size(test_font_size)
            
            all_processed_lines = []
            total_height = 0
            fits_in_one_line = True
            
            for item in text_lines:
                # Обрабатываем заголовок и значение отдельно
                safe_header = _force_wrap_text(pdf, item['header'], max_header_width)
                safe_value = _force_wrap_text(pdf, item['value'], value_width)
                
                # Проверяем, помещается ли текст в одну строку
                pdf.set_font(font_family, 'B')
                header_fits = text_fits_in_one_line(test_font_size, safe_header, max_header_width)
                
                pdf.set_font(font_family, '')
                value_fits = text_fits_in_one_line(test_font_size, safe_value, value_width)
                
                if not (header_fits and value_fits):
                    fits_in_one_line = False
                
                # Добавляем обработанные строки как словарь
                all_processed_lines.append({"header": safe_header, "value": safe_value})

            # Calculate height based on the safe, wrapped lines
            for item in all_processed_lines:
                # Считаем высоту для заголовка
                pdf.set_font(font_family, 'B')
                header_lines = len(pdf.multi_cell(w=max_header_width, txt=item['header'], split_only=True))
                
                # Считаем высоту для значения
                pdf.set_font(font_family, '')
                value_lines = len(pdf.multi_cell(w=value_width, txt=item['value'], split_only=True))
                
                # Берем максимальное количество строк (заголовок или значение)
                max_lines = max(header_lines, value_lines)
                
                # Суммируем высоту (используем максимальное количество строк, так как они будут отображаться рядом)
                total_height += max_lines * pdf.font_size + 2  # +2 для отступа между строками таблицы
            
            # Проверяем, что текст помещается по высоте и все строки помещаются в одну строку
            if total_height < available_height and fits_in_one_line:
                best_font_size = test_font_size
                final_lines_to_render = all_processed_lines
                break 

        # Если не удалось найти размер шрифта, при котором все строки помещаются в одну строку,
        # выбираем наименьший размер шрифта
        if best_font_size == 0:
            best_font_size = 6
            pdf.set_font_size(best_font_size)
            # Process with the smallest font size
            all_processed_lines_fallback = []
            for line in text_lines:
                # Обрабатываем заголовок и значение отдельно
                safe_header = _force_wrap_text(pdf, line['header'], max_header_width)
                safe_value = _force_wrap_text(pdf, line['value'], value_width)
                
                # Добавляем обработанные строки как словарь
                all_processed_lines_fallback.append({'header': safe_header, 'value': safe_value})
            final_lines_to_render = all_processed_lines_fallback
            
            # Optional: check height again for warning
            total_height_fallback = 0
            for item in final_lines_to_render:
                # Считаем высоту для заголовка
                pdf.set_font(font_family, 'B')
                header_lines = len(pdf.multi_cell(w=max_header_width, txt=item['header'], split_only=True))
                
                # Считаем высоту для значения
                pdf.set_font(font_family, '')
                # Для значений разрешаем перенос по словам
                value_lines = len(pdf.multi_cell(w=value_width, txt=item['value'], split_only=True))
                
                # Берем максимальное количество строк (заголовок или значение)
                max_lines = max(header_lines, value_lines)
                
                # Суммируем высоту (используем максимальное количество строк, так как они будут отображаться рядом)
                total_height_fallback += max_lines * pdf.font_size + 1  # +1 для отступа между строками таблицы
            if total_height_fallback > available_height:
                 logger.warning(f"Текст для артикула {article} не помещается по высоте даже с минимальным шрифтом. Возможны искажения.")
            if not all(text_fits_in_one_line(best_font_size, item['header'], max_header_width) 
                       for item in final_lines_to_render):
                 logger.warning(f"Некоторые заголовки для артикула {article} не помещаются в одну строку даже с минимальным шрифтом.")

        pdf.set_font_size(best_font_size)

        # Добавляем каждую строку текста в PDF в формате двухколоночной таблицы
        for item in final_lines_to_render:
            # Сохраняем текущую позицию X и Y
            x_pos = pdf.get_x()
            y_pos = pdf.get_y()
            
            # Устанавливаем жирный шрифт для заголовка (левая колонка)
            pdf.set_font_size(best_font_size)
            pdf.set_font(font_family, 'B')  # B - жирный шрифт
            
            # Проверяем, помещается ли заголовок в одну строку
            header_fits = text_fits_in_one_line(best_font_size, item['header'], max_header_width)
            
            # Заголовок всегда должен быть в одну строку
            pdf.set_font(font_family, 'B')
            header_text = item['header']
            if not header_fits:
                # Обрезаем текст, чтобы он поместился в одну строку
                truncated_header = ""
                for char in header_text:
                    if pdf.get_string_width(truncated_header + char + "...") <= max_header_width:
                        truncated_header += char
                    else:
                        break
                header_text = truncated_header + "..."
            
            # Отрисовываем заголовок (левая колонка)
            pdf.set_xy(x_pos, y_pos)
            pdf.set_font(font_family, 'B')
            pdf.cell(w=max_header_width, h=pdf.font_size, txt=header_text, align='L')
            
            # Для значения разрешаем перенос по словам
            pdf.set_font(font_family, '')
            value_text = item['value']
            
            # Отрисовываем значение (правая колонка) с переносом по словам
            pdf.set_xy(x_pos + max_header_width + 10, y_pos)  # +10 для отступа между колонками
            
            # Используем multi_cell для значения, чтобы разрешить перенос по словам
            value_lines = pdf.multi_cell(w=value_width, h=pdf.font_size, txt=value_text, align='L', split_only=True)
            
            # Если значение помещается в одну строку, используем cell для лучшего выравнивания
            if len(value_lines) == 1:
                pdf.set_xy(x_pos + max_header_width + 10, y_pos)
                pdf.cell(w=value_width, h=pdf.font_size, txt=value_text, align='L')
                # Перемещаемся на следующую строку
                pdf.set_y(y_pos + pdf.font_size + 2)  # +2 для отступа между строками
            else:
                # Если значение не помещается в одну строку, используем multi_cell
                pdf.set_xy(x_pos + max_header_width + 10, y_pos)
                line_height = pdf.font_size
                pdf.multi_cell(w=value_width, h=line_height, txt=value_text, align='L')
                # Перемещаемся на следующую строку после multi_cell
                # multi_cell автоматически перемещает курсор вниз, поэтому нам нужно только добавить отступ
                pdf.set_y(pdf.get_y() + 2)  # +2 для отступа между строками
            
        inserted_cards += 1

    if inserted_cards == 0:
        return "", 0, not_found_articles

    output_filename = f"product_cards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    output_path = os.path.join(output_folder, output_filename)
    pdf.output(output_path)
    
    return output_path, inserted_cards, not_found_articles
